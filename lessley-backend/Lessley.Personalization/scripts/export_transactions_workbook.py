"""
Export a transaction feed to Excel, with every insight's own arithmetic shown per row.

The point is validation, not reporting. Each query gets its own column, holding what that one
transaction contributes to that one query's answer — so the user can sum a column in Excel and
land on the number the app shows, or find the rows where it does not.

Nothing here re-implements a calculation. Every figure comes from the same `TransactionAmountService`
and `InsightsService` the API calls, against the same reference data in MongoDB, so a disagreement
between a column and the screen is a real disagreement rather than two guesses at the same rule.

Usage
    python scripts/export_transactions_workbook.py ../../local-data/transactions-roee-365-25082026.json \
        --email roeecr@gmail.com \
        --out ../../local-data/roee-validation.xlsx

A club card needs no flag and no second API call: a purchase on a loaded נטען card arrives as a
settled row the card was never billed for, and that signature is in the feed already.
"""

import argparse
import asyncio
import json
import os
import sys
from collections import defaultdict

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from database.db_client import init_db, close_db
from models.transaction import Transaction
from services.insights_service import InsightsService
from services.reference_data_repository import ReferenceDataRepository
from services.transaction_amount_service import TransactionAmountService

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(color="FFFFFF", bold=True)
GROUP_FILL = PatternFill("solid", fgColor="D9E1F2")
MONEY = "#,##0.00"


def load_transactions(path: str) -> list[Transaction]:
    """The feed, whichever wrapper it arrived in — devtools and the API disagree."""
    with open(path, encoding="utf-8") as handle:
        payload = json.load(handle)
    rows = payload if isinstance(payload, list) else (payload.get("items") or payload.get("data") or [])
    return [Transaction.model_validate(row) for row in rows]


def date_of(transaction: Transaction):
    if not transaction.date:
        return None
    return transaction.date.transactionDate or transaction.date.bookingDate or transaction.date.valueDate


def currency_of(detail) -> str:
    return (detail.currency if detail else None) or ""


def amount_of(detail):
    return detail.amount if detail else None


def style_header(sheet, row: int = 1) -> None:
    for cell in sheet[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    # By reference, not by `sheet.cell(...)`: fetching a cell materialises it, which moves
    # `max_row` past the header and makes the next `append` skip a row — silently shifting
    # every data row down one and leaving the SUM ranges below pointing a row short.
    sheet.freeze_panes = f"A{row + 1}"


def autosize(sheet, widths: dict[int, int]) -> None:
    for index, width in widths.items():
        sheet.column_dimensions[get_column_letter(index)].width = width


async def build(args) -> int:
    transactions = load_transactions(args.transactions)
    if not transactions:
        print(f"No transactions found in {args.transactions}")
        return 1

    await init_db()
    reference = ReferenceDataRepository()
    await reference.load_async()

    amounts = TransactionAmountService()
    insights = InsightsService(
        open_finance_service=None,
        files_service=None,
        publisher_service=None,
        user_repository=None,
        reference_data_repository=reference,
        mcc_service=None,
        amount_service=amounts,
    )

    club_ids = args.clubs or await user_clubs(args.email)
    print(f"{len(transactions)} transactions · clubs: {club_ids}")

    # ── The savings answer, exactly as the endpoint returns it ───────────────
    #
    # Sorted first, because the orchestrator does.
    ordered = sorted(transactions, key=lambda t: (date_of(t) is None, date_of(t)), reverse=True)
    answer = insights.savings_opportunities(ordered, user_club_ids=club_ids)

    # Every purchase appears under exactly one band, and never on both sides, so these are
    # plain lookups rather than the reconciliation this script used to have to do.
    band_of: dict[str, str] = {}
    merchant_shops: dict[str, list[str]] = {}
    for band in answer.missed.bands:
        for merchant in band.merchants:
            names = [shop.store_name for shop in merchant.shops]
            for purchase in merchant.purchases:
                band_of[purchase.transaction_id] = band.band
                merchant_shops[purchase.transaction_id] = names

    missed_ids = set(band_of)
    applied_ids = {
        purchase.transaction_id
        for merchant in answer.applied.merchants
        for purchase in merchant.purchases
    }

    workbook = Workbook()
    write_transactions(workbook, transactions, amounts, merchant_shops, band_of, missed_ids, applied_ids)
    write_totals(workbook, insights, amounts, transactions, answer, len(transactions))
    write_savings(workbook, answer)
    write_by_account(workbook, insights, transactions)
    write_overview(workbook, insights, transactions)

    workbook.save(args.out)
    print(f"\nWrote {args.out}")
    return 0


async def user_clubs(email: str | None) -> list[str] | None:
    """The user's memberships, so the matching drops deals they cannot use."""
    if not email:
        return None
    from models.db.entities import Club  # noqa: F401  (registers the model with Beanie)
    from config.settings import settings
    from pymongo import MongoClient

    client = MongoClient(settings.ConnectionStrings_MongoDb)
    user = client["lessley"]["users"].find_one({"NormalizedEmail": email.upper()})
    client.close()
    return (user or {}).get("Clubs")


def write_transactions(workbook, transactions, amounts, merchant_shops, band_of, missed_ids, applied_ids):
    """One row per transaction: the raw amounts, then what each query takes from it."""
    sheet = workbook.active
    sheet.title = "Transactions"

    columns = [
        ("id", 26), ("date", 11), ("merchant", 30), ("category", 18),
        ("accountId", 20), ("accountNumber", 20), ("provider", 10), ("status", 9),
        ("original", 12), ("orig ccy", 9), ("charged", 12), ("chg ccy", 9),
        ("markup fee", 11), ("installment?", 11), ("duplicate?", 10),
        ("kind", 12), ("direction", 10), ("is purchase", 10),
        ("SAVED", 11), ("saving reason", 20),
        ("MISSED", 11), ("APPLIED", 11), ("matched shops", 34), ("band", 10), ("club card?", 10),
        ("q: total-amount", 14), ("q: save-with-lessley", 14), ("q: spending-by-accounts", 15),
        ("q: top-accounts", 14), ("q: missed-discount", 14), ("q: discount-applied", 14),
        ("q: spending-overview", 14), ("overview kind", 13),
    ]
    sheet.append([name for name, _ in columns])
    style_header(sheet)
    autosize(sheet, {i + 1: width for i, (_, width) in enumerate(columns)})

    for transaction in transactions:
        saving, reason = amounts.saving_of(transaction)
        spent = amounts.amount_spent(transaction)
        received = amounts.amount_received(transaction)
        value = amounts.purchase_value(transaction)
        kind = amounts.kind_of(transaction)
        tid = transaction.id or ""

        is_missed = tid in missed_ids
        is_applied = tid in applied_ids

        sheet.append([
            tid,
            date_of(transaction),
            transaction.merchantName or "",
            (transaction.category.sub if transaction.category else "") or "",
            (transaction.accountId or "")[-12:],
            transaction.accountNumber or "",
            transaction.providerId or "",
            transaction.status or "",
            amount_of(transaction.amount.originalAmount if transaction.amount else None),
            currency_of(transaction.amount.originalAmount if transaction.amount else None),
            amount_of(transaction.amount.chargedAmount if transaction.amount else None),
            currency_of(transaction.amount.chargedAmount if transaction.amount else None),
            amounts.markup_fee(transaction),
            "yes" if transaction.isCreditCardInstallment else "",
            "yes" if transaction.isDuplicate else "",
            kind,
            amounts.direction(transaction),
            "yes" if amounts.is_purchase(transaction) else "",
            saving,
            reason,
            value if is_missed else 0.0,
            value if is_applied else 0.0,
            ", ".join(sorted(set(merchant_shops.get(tid, [])))[:4]),
            # One band, never a list: a purchase is counted under its strongest match only.
            band_of.get(tid, ""),
            "yes" if amounts.paid_with_club_card(transaction) else "",
            # ── what each query takes from this row ───────────────────────────
            amounts.net_purchase_value(transaction),             # spending-total: total_amount
            saving,                                              # spending-saved
            saving,                                              # spending-saved-by-account (group by accountId)
            amounts.net_purchase_value(transaction),             # top-accounts (group by accountId)
            value if is_missed else 0.0,                         # missed-savings header
            value if is_applied else 0.0,                        # discounts-applied header
            amounts.net_purchase_value(transaction),             # spending overview (signed, per composition)
            kind,
        ])

    money_columns = [9, 11, 13, 19, 21, 22, 26, 27, 28, 29, 30, 31, 32]
    for row in sheet.iter_rows(min_row=2):
        for index in money_columns:
            row[index - 1].number_format = MONEY
        row[1].number_format = "yyyy-mm-dd"


def write_totals(workbook, insights, amounts, transactions, answer, row_count):
    """
    Each query's own answer, beside a live SUM of its column — so Excel does the checking.

    The `difference` column is the whole sheet. Anything but zero there is a genuine
    disagreement between what the service computes and what the per-row breakdown claims.
    """
    sheet = workbook.create_sheet("Query totals")
    sheet.append(["query", "endpoint", "service says", "sum of column", "difference", "counts", "what it sums"])
    style_header(sheet)
    autosize(sheet, {1: 24, 2: 34, 3: 16, 4: 16, 5: 12, 6: 10, 7: 62})

    countable = amounts.countable(transactions)
    total = insights.spending_total(transactions)
    saved = insights.spending_saved(transactions)
    last = row_count + 1

    def column(letter: str) -> str:
        return f"=SUM(Transactions!{letter}2:{letter}{last})"

    rows = [
        ("total-amount", "/insights/spending-total", total["total_amount"], column("Z"),
         total["purchase_count"],
         "net_purchase_value: everything bought less what came back. A club-card purchase counts at full price."),
        ("save-with-lessley", "/insights/spending-saved", saved, column("AA"),
         sum(1 for t in transactions if amounts.amount_saved(t)),
         "amount_saved: the original-vs-charged gap, plus club-card purchases, plus refunds."),
        ("spending-by-accounts", "/insights/spending-saved-by-account", saved, column("AB"),
         len({t.accountId for t in transactions if t.accountId}),
         "the same amount_saved, grouped by accountId. Pivot column AB by column E to check each account."),
        ("top-accounts", "/insights/top-accounts", sum(amounts.net_purchase_value(t) for t in countable),
         column("AC"), len({t.accountId for t in countable if t.accountId}),
         "net_purchase_value grouped by accountId. Now the same figure as total-amount, so the two agree."),
        ("missed-discount", "/insights/savings-opportunities (missed)", answer.missed.total_amount,
         column("AD"), answer.missed.purchase_count,
         "purchase_value of every purchase that matched a deal-running shop and was NOT paid with a club card. "
         "Counted once, under its strongest band."),
        ("discount-applied", "/insights/savings-opportunities (applied)", answer.applied.total_amount,
         column("AE"), answer.applied.purchase_count,
         "purchase_value of every purchase a club card paid for, whether or not our catalogue knows the shop."),
        ("spending-overview", "/insights/spending-total (composition)", sum(
            row["signed_amount"] for row in total["composition"]), column("AF"), len(countable),
         "signed_amount per kind, which sums to total-amount exactly. See the Overview sheet."),
    ]

    for name, endpoint, service_value, formula, count, explanation in rows:
        sheet.append([name, endpoint, service_value, formula, None, count, explanation])

    for index, row in enumerate(sheet.iter_rows(min_row=2), start=2):
        row[2].number_format = MONEY
        row[3].number_format = MONEY
        row[4].value = f"=ROUND(C{index}-D{index}, 2)"
        row[4].number_format = MONEY
        row[6].alignment = Alignment(wrap_text=True, vertical="top")

    note = sheet.cell(row=len(rows) + 3, column=1)
    note.value = (
        "The band tabs on screen now DO sum to missed-discount: every purchase is counted under "
        "its strongest match and nowhere else. The 'Savings' sheet lists the bands with their own "
        "subtotals so that can be checked directly."
    )
    note.alignment = Alignment(wrap_text=True, vertical="top")
    sheet.merge_cells(start_row=note.row, start_column=1, end_row=note.row + 2, end_column=7)


def write_savings(workbook, answer):
    """
    The savings payload itself: one row per merchant, with the band it was counted under.

    The band subtotals are printed beside a live SUM of the merchants under them, so the
    partition can be checked without leaving the sheet.
    """
    sheet = workbook.create_sheet("Savings")
    sheet.append(["side", "band", "merchant", "purchases", "amount", "deals", "clubs", "shops", "transaction ids"])
    style_header(sheet)
    autosize(sheet, {1: 9, 2: 9, 3: 32, 4: 10, 5: 12, 6: 7, 7: 30, 8: 40, 9: 52})

    for band in answer.missed.bands:
        for merchant in band.merchants:
            sheet.append([
                "missed", band.band, merchant.merchant_name, merchant.purchase_count, merchant.amount,
                merchant.deal_count, ", ".join(merchant.club_ids),
                ", ".join(shop.store_name for shop in merchant.shops),
                ", ".join(p.transaction_id for p in merchant.purchases),
            ])
    for merchant in answer.applied.merchants:
        sheet.append([
            "applied", "", merchant.merchant_name, merchant.purchase_count, merchant.amount,
            merchant.deal_count, ", ".join(merchant.club_ids),
            ", ".join(shop.store_name for shop in merchant.shops),
            ", ".join(p.transaction_id for p in merchant.purchases),
        ])
    for row in sheet.iter_rows(min_row=2):
        row[4].number_format = MONEY

    first_row = sheet.max_row + 2
    sheet.cell(row=first_row, column=1).value = "band"
    sheet.cell(row=first_row, column=2).value = "service says"
    sheet.cell(row=first_row, column=3).value = "sum of its merchants"
    for offset, band in enumerate(answer.missed.bands, start=1):
        row = first_row + offset
        sheet.cell(row=row, column=1).value = band.band
        sheet.cell(row=row, column=2).value = band.total_amount
        sheet.cell(row=row, column=3).value = (
            f'=SUMIFS(E2:E{first_row - 2}, A2:A{first_row - 2}, "missed", B2:B{first_row - 2}, "{band.band}")'
        )
        sheet.cell(row=row, column=2).number_format = MONEY
        sheet.cell(row=row, column=3).number_format = MONEY

    total_row = first_row + len(answer.missed.bands) + 1
    sheet.cell(row=total_row, column=1).value = "missed total"
    sheet.cell(row=total_row, column=2).value = answer.missed.total_amount
    sheet.cell(row=total_row, column=3).value = (
        f"=SUM(C{first_row + 1}:C{first_row + len(answer.missed.bands)})"
    )
    sheet.cell(row=total_row, column=2).number_format = MONEY
    sheet.cell(row=total_row, column=3).number_format = MONEY


def write_by_account(workbook, insights, transactions):
    """Per-account savings and spend, the two the endpoints answer separately."""
    sheet = workbook.create_sheet("By account")
    sheet.append(["accountId", "accountNumber", "saved (spending-saved-by-account)",
                  "purchases (top-accounts)", "spend (top-accounts)"])
    style_header(sheet)
    autosize(sheet, {1: 20, 2: 20, 3: 26, 4: 20, 5: 20})

    saved_rows = {row["accountId"]: row for row in insights.spending_saved_by_account(transactions)}
    for account in insights.top_spending_accounts(transactions, limit=100):
        saved = saved_rows.get(account["accountId"], {})
        sheet.append([
            (account["accountId"] or "")[-12:], account["accountNumber"],
            saved.get("total_saved", 0.0), account["total_count"], account["total_amount"],
        ])
    for row in sheet.iter_rows(min_row=2):
        row[2].number_format = MONEY
        row[4].number_format = MONEY


def write_overview(workbook, insights, transactions):
    """The composition behind spending-overview: what the period was actually made of."""
    sheet = workbook.create_sheet("Overview")
    total = insights.spending_total(transactions)
    sheet.append(["kind", "count", "amount (shown)", "signed (counted)", "markup fees", "distinct plans"])
    style_header(sheet)
    autosize(sheet, {1: 16, 2: 9, 3: 16, 4: 16, 5: 14, 6: 14})

    for row in total["composition"]:
        sheet.append([row["kind"], row["count"], row["amount"], row["signed_amount"],
                      row.get("markup_fees"), row.get("plan_count")])
    for row in sheet.iter_rows(min_row=2):
        for index in (3, 4, 5):
            row[index - 1].number_format = MONEY

    last = len(total["composition"]) + 1
    sheet.append([])
    sheet.append(["total-amount (spending-total)", total["purchase_count"], None, total["total_amount"]])
    sheet.append(["sum of the signed column", None, None, f"=SUM(D2:D{last})"])
    sheet.append([
        "'amount' is what the screen prints \u2014 a refund reads as what came back, not as a negative. "
        "'signed' is the same money by which way it moved, and it is that column which sums to "
        "total-amount exactly."
    ])
    for row in sheet.iter_rows(min_row=last + 2):
        row[3].number_format = MONEY
    sheet.cell(row=last + 4, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    sheet.merge_cells(start_row=last + 4, start_column=1, end_row=last + 5, end_column=5)


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("transactions", help="the transaction feed, as saved from the API")
    parser.add_argument("--out", default="transactions-validation.xlsx")
    parser.add_argument("--email", help="whose clubs to match against, e.g. roeecr@gmail.com")
    parser.add_argument("--clubs", nargs="*", help="club ids, overriding --email")
    args = parser.parse_args()

    try:
        return asyncio.run(build(args))
    finally:
        asyncio.run(close_db())


if __name__ == "__main__":
    raise SystemExit(main())
